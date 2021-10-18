#! C:\Users\bluem\Documents\Sandbox\python\venvs\automation\Scripts\python.exe
'''
To-do:
'''

from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import subprocess
from pathlib import Path
import smtplib, ssl


x = Path(r'X:')

if not x.exists():
    subprocess.call(r'net use X: \\server1\admin /user:Administrator Server1', shell=True)

# file = 'C:/Users/bluem/Documents/Sandbox/python/automate_the_boring_stuff_with_python/excel/datum/datum.xlsx'
file = 'X:/Documents/datum.xlsx'
wb = load_workbook(file, data_only = True)

col_prohlidka = ord('D') - 64
col_bozp = ord('H') - 64

today = date.today()

ws = wb['Sheet1']

sender = 'pyzdenek@gmail.com'
reciever = 'habahaba0123@gmail.com'
password = 'Ma9G9#JZ'
port = 465

message = MIMEMultipart('alternative')
message['Subject'] = 'kontrola terminu'
message['From'] = sender
message['To'] = reciever

body = ''

for i in range(3,100):
    platnost_prohlidka = (ws.cell(row=i,column=col_prohlidka).value, col_prohlidka, 'zdr. prohlidka:')
    platnost_bozp = (ws.cell(row=i,column=col_bozp).value, col_bozp, 'bozp:')
    platnost = [platnost_prohlidka, platnost_bozp]

    for p in platnost:
        if p[0]:
            if ws.cell(row=i, column=p[1]+2).value == 0:
                continue
            else:
                c = ws.cell(row=i, column=p[1]).value.date()
                deadline = c - timedelta(days = 14)
                # print(f'[{i}] {deadline}')
                if deadline <= today:
                    jmeno, dn = ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value.date()
                    report = f'[{str(i).zfill(2)}] Jmeno: <b>{jmeno:<10}</b>| narozen/a: <b>{dn.strftime("%d.%m.%Y")}</b>| <b>{p[2]:<14}</b> plati do <b>{c.strftime("%d.%m.%Y")}</b><br>'
                    body += report


print(body)
if body != '':
    message.attach(MIMEText(body,'html'))

    # Create a secure SSL context
    context = ssl.create_default_context()

    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.login(sender, password)
        server.sendmail(sender, reciever, message.as_string())




