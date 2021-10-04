from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
from random import randint

file = 'datum.xlsx'
wb = load_workbook(file)
wb.iso_dates = True
ws = wb['Sheet1']


# for i in range(2,51):
#     year = (str(randint(1964,1990)), str(randint(2019,2027)), str(randint(2019,2027)))
#     # year = randint(1965,1990)
#     month = (str(randint(1,12)).zfill(2), str(randint(1,12)).zfill(2), str(randint(1,12)).zfill(2))
#     day = []

#     for m in month:
#         if m in [4,6,9,11]:
#             day.append(str(randint(1,30)).zfill(2))
#         elif m == 2:
#             day.append(str(randint(1,28)).zfill(2))
#         else:
#             day.append(str(randint(1,31)).zfill(2))

#     datum1 = f'{str(day[0])}.{str(month[0])}.{str(year[0])}'
#     # ws['B'+str(i)].value = datum1
#     ws.cell(column=2,row=i).value=datum1
#     ws.cell(column=2,row=i).number_format = 'dd.mm.yyyy;@'

#     datum2 = f'{str(day[1])}.{str(month[1])}.{str(year[1])}'
#     # ws['D'+str(i)].value = datum2
#     ws.cell(column=4,row=i).value=datum2
#     ws.cell(column=4,row=i).number_format = 'dd.mm.yyyy;@'

#     datum3 = f'{str(day[2])}.{str(month[2])}.{str(year[2])}'
#     # ws['F'+str(i)].value = datum2
#     ws.cell(column=6,row=i).value=datum3
#     ws.cell(column=6,row=i).number_format = 'dd.mm.yyyy;@'

for i in range(3,51):
    datum = ws.cell(column=3,row=i).value
    # print(datum)
    # print(datum.year)
    newDatum = f'{randint(2016,2020)}-{datum.month}-{datum.day} 00:00:00'
    ws.cell(column=3,row=i).value = newDatum



wb.save(file)
