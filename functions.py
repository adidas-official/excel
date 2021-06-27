from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from unidecode import unidecode
import unicodedata

def strip_accents(text):

    try:
        text = unicode(text, 'utf-8')
    except NameError: # unicode is a default on python 3
        pass

    text = unicodedata.normalize('NFD', text)\
           .encode('ascii', 'ignore')\
           .decode("utf-8")

    return str(text)


def completeArea(area):
# Doplneni na obdelnikovou oblast.
# Input=list(leva-horni, prava-dolni)
# Priklad:
#   completeArea(['A3', 'D4']) -> ['A3','B3','C3','D3','A4','B4','C4','D4']
# Output=list

    completeArea = []
    firstCol = ws[area[0]].column
    lastCol = ws[area[1]].column
    firstRow = ws[area[0]].row
    lastRow = ws[area[1]].row
    # print(firstCol, lastCol) # 1 3
    # print(firstRow, lastRow) # 2 5

    for col in range(firstCol, lastCol + 1):
        for row in range(firstRow, lastRow + 1):
            # cell = str(col) + str(row)
            cell = ws.cell(column=col, row=row).coordinate
            completeArea.append(cell)
    return completeArea


def findCellInArea(subject, area):
# Vyhleda 'subject' ve vybrane oblasti 'area'. Input: subject=string, area=list
# Priklad:
#   findCellInArea('Karel', ['A2', 'C9'])

    for cell in completeArea(area):
        if ws[cell].value == subject:
            return cell

    return False


def rady(depth):
# Funkce pro pocitani lidi v seznamu, Input=int, Output=int
# Prohleda jen prvni sloupec od treti rady dolu
    pocetLidi = 0
    for row in range(3,depth):
        cell = ws["A"+str(row)].value
        if cell:
            pocetLidi += 1

    return pocetLidi


def formatTxt(names):
    lines = [i for i, s in enumerate(names) if '-------' in s]

    s_lines = sorted(lines, reverse=True)

    for index in s_lines:
        if index == s_lines[-1]: # last element
            del names[:lines[0]+1]
        elif index == s_lines[0]: # first element
            del names[lines[-1]:]
        else:
            del names[index-1:index+1]

    data = {}
    for name in names:

        n = " ".join(name.split())
        splited = n.split(' ')
        mzda, jmeno = splited[-1].replace('.',''), splited[0]

        if '/' in mzda:
            mzda = 0
        else:
            mzda = int(mzda)

        data.setdefault(jmeno,[])
        data[jmeno].append(mzda)
    return data


def findMonth(ws, month, width, depth=100):
    for column in range(2,depth,width):
        m = strip_accents(ws.cell(column=column,row=1).value).lower()
        if m == month:
            return get_column_letter(column)
            # return column
    return False


