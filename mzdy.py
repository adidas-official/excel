'''
To-do:
- Formatovani
- kopie tabulky

'''

import functions
import io
import msoffcrypto
import os
from shutil import copyfile
from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
from csv import reader
from sys import argv


file = "Mzdové náklady 2021.xlsx"
os.system('cp '+file+' temp.xlsx')
decrypted_wb = io.BytesIO()

with open(file, 'rb') as f:
    officeFile = msoffcrypto.OfficeFile(f)
    officeFile.load_key(password='13881744')
    officeFile.decrypt(decrypted_wb)

wb = load_workbook(filename=decrypted_wb)


# Vyhleda mesic v prvnim radku s defaultni hloubkou depth=100
def findMonth(month, width, depth=100):
    for column in range(2,depth,width):
        m = functions.strip_accents(ws.cell(column=column,row=1).value).lower()
        if m == month:
            return get_column_letter(column)
            # return column
    return False


def formatTxt(names):
    lines = [i for i, s in enumerate(names) if '-------' in s]

    s_lines = sorted(lines, reverse=True)

    for index in s_lines:
        if index == s_lines[-1]: # last element
            del names[:lines[0]+1]
        elif index == s_lines[0]: # first element
            del names[lines[-1]:]
        else:
            del names[index-1:index+1]

    data = {}
    for name in names:

        n = " ".join(name.split())
        splited = n.split(' ')
        mzda, jmeno = splited[-1].replace('.',''), splited[0]

        if '/' in mzda:
            mzda = 0
        else:
            mzda = int(mzda)

        data.setdefault(jmeno,[])
        data[jmeno].append(mzda)
    return data


# updateFile = 'TEXT.TXT'
# file = "Mzdové náklady 2021-clean.xlsx"
# wb = load_workbook(file)
updateFile = argv[2]

with open(updateFile,'r',encoding='windows-1250') as f:
    width = 6
    month = str(argv[1])
    print(month)
    names = f.readlines()
    data = formatTxt(names)

    seznamJmen = {}

    for worksheet in wb.worksheets:
        if worksheet.title not in ['Celkový součet', 'kontrola', 'List1']:
            if worksheet.title in ['Úřad práce', 'Úřad práce Cheb', 'Katastrální úřad']:
                width = 4
            else:
                width = 6
            ws = wb[worksheet.title]

            mesic = findMonth(month,width)

            for i in range(1,200):
                bunka = ws.cell(row = i, column = 1)
                cele_jmeno = bunka.value
                if cele_jmeno == 'Zákonné pojištění' or cele_jmeno == 'Mzdové náklady': # Konec jmen
                    break
                if cele_jmeno:
                    cele_jmeno = cele_jmeno.split(' ')
                    if len(cele_jmeno) > 1:
                        prijmeni = cele_jmeno[0]
                        seznamJmen.setdefault(prijmeni,{})
                        seznamJmen[prijmeni].setdefault(worksheet.title,[])
                        seznamJmen[prijmeni][worksheet.title].append(mesic+str(bunka.row))

    cols = (25,25,10)
    for jmeno in seznamJmen:
        counter = 0
        if jmeno in data:
            for k,v in seznamJmen[jmeno].items():
                for bunka in v:
                    if counter < len(data[jmeno]):
                        print('+'+'-' * 65 + '+')
                        print('| ' + str(jmeno).ljust(cols[0]) + '| ' + str(k).ljust(cols[1]) + '| ' + str(data[jmeno][counter]).ljust(cols[2]) + '|')
                        ws = wb[k]
                        ws[bunka].value = data[jmeno][counter]
                        counter += 1

    print('+'+'-' * 65 + '+')
    print()
    print('!!! NEZAPOMEN ZA SEBOU ZAMKNOUT !!!')
wb.save(file)

