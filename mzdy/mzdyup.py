# coding=Windows-1250

'''
TODO:
open xls
copy status invalidy
choose month
'''

from time import process_time
t0 = process_time()

import functions, msoffcrypto
from os import system
from win32com.client import Dispatch
from shutil import copyfile
from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
# from sys import argv
# month = argv[1]
mesic = 'P'

updateFile = 'TEXT.TXT'
file = "jmenny_seznam_2021_10_01 Bereko.xlsx"
tempFile = 'temp.xlsx'
# copyfile(file,tempFile)
# xl = Dispatch('Excel.Application')
# bw = SaveAs('tempFile.xlsx', FileFormat=51)
# xl.Quit()


with open(updateFile,'r',encoding='windows-1250') as f:
    names = f.readlines()
    data = functions.formatTxt(names) # 'Běhůnková': [11045], 'Bobok': [412], ... TXT

seznamJmen = {}
missing = []

wb = load_workbook(tempFile)
ws = wb['2) jmenný seznam']

for i in range(1,200):
    bunka = ws.cell(row = i, column = 2)
    cele_jmeno = bunka.value

    if cele_jmeno:
        seznamJmen.setdefault(cele_jmeno,[])
        seznamJmen[cele_jmeno].append(mesic+str(bunka.row)) #'Mihalík': ['AZ21'], 'Musilová': ['AZ22'],... EXCEL

cols = (25,25,10)
print()
print('VYPLNUJI MZDU U OSOB:')
for jmeno in seznamJmen:
    counter = 0
    if jmeno in data:
        for bunka in seznamJmen[jmeno]:
            if counter < len(data[jmeno]):
                print('+'+'-' * 38 + '+')
                print('| ' + str(jmeno).ljust(cols[0]) + '| ' + str(data[jmeno][counter]).ljust(cols[2]) + '|')
                ws[bunka].value = data[jmeno][counter]
                counter += 1

print('+'+'-' * 38 + '+')

print('\nCHYBEJICI JMENA V TABULCE:')
for name,pay in data.items():
    if name not in seznamJmen:
        for i in pay:
            print('+'+'-' * 38 + '+')
            print('| ' + str(name).ljust(cols[0]) + '| ' + str(i).ljust(cols[2]) + '|')

print('+'+'-' * 38 + '+')
# print()

# print('!!! NEZAPOMEN ZA SEBOU ZAMKNOUT !!!')

# wb.save(tempFile)

t1 = process_time()
print(f'Time processed: {t1 - t0}s')

# system(tempFile)
