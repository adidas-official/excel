'''
To-do:

'''

from time import process_time
t0 = process_time()

import functions, msoffcrypto
from os import system
from io import BytesIO
from shutil import copyfile
from openpyxl import Workbook, load_workbook, workbook
from openpyxl.utils import get_column_letter
from sys import argv


month = argv[1]
updateFile = argv[2]

file = "Mzdové náklady 2021.xlsx"
tempFile = 'temp.xlsx'
copyfile(file,tempFile)

try:
    decrypted_wb = BytesIO()
    with open(tempFile, 'rb') as f:
        officeFile = msoffcrypto.OfficeFile(f)
        officeFile.load_key(password='13881744')
        officeFile.decrypt(decrypted_wb)

    wb = load_workbook(filename=decrypted_wb)
except UnboundLocalError:
    wb = load_workbook(tempFile)


with open(updateFile,'r',encoding='windows-1250') as f:
    names = f.readlines()
    data = functions.formatTxt(names)

width = 6
seznamJmen = {}

for worksheet in wb.worksheets:
    if worksheet.title not in ['Celkový součet', 'kontrola', 'List1']:
        if worksheet.title in ['Úřad práce', 'Úřad práce Cheb', 'Katastrální úřad']:
            width = 4
        else:
            width = 6
        ws = wb[worksheet.title]

        mesic = functions.findMonth(ws, month,width)

        for i in range(1,200):
            bunka = ws.cell(row = i, column = 1)
            cele_jmeno = bunka.value
            if cele_jmeno == 'Zákonné pojištění' or cele_jmeno == 'Mzdové náklady': # Konec jmen
                break
            if cele_jmeno:
                cele_jmeno = cele_jmeno.split(' ')
                if len(cele_jmeno) > 1:
                    prijmeni = cele_jmeno[0]
                    seznamJmen.setdefault(prijmeni,{})
                    seznamJmen[prijmeni].setdefault(worksheet.title,[])
                    seznamJmen[prijmeni][worksheet.title].append(mesic+str(bunka.row))

cols = (25,25,10)
for jmeno in seznamJmen:
    counter = 0
    if jmeno in data:
        for k,v in seznamJmen[jmeno].items():
            for bunka in v:
                if counter < len(data[jmeno]):
                    print('+'+'-' * 65 + '+')
                    print('| ' + str(jmeno).ljust(cols[0]) + '| ' + str(k).ljust(cols[1]) + '| ' + str(data[jmeno][counter]).ljust(cols[2]) + '|')
                    ws = wb[k]
                    ws[bunka].value = data[jmeno][counter]
                    counter += 1

print('+'+'-' * 65 + '+')
print()
print('!!! NEZAPOMEN ZA SEBOU ZAMKNOUT !!!')

wb.save(tempFile)

t1 = process_time()
print(f'Time processed: {t1 - t0}s')

system(tempFile)

